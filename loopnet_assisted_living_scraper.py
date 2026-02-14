"""
LoopNet Healthcare — Assisted Living Facility Scraper v2
=========================================================
Scrapes multiple U.S. states for assisted living facilities listed FOR SALE
under the Healthcare > Assisted Living category on LoopNet.
Output file includes timestamp so it never conflicts with an open file.

SETUP (run once):
    pip install playwright openpyxl
    python -m playwright install chromium

USAGE:
    python loopnet_assisted_living_scraper.py
"""

import time
import re
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Pre-compiled regex patterns (avoids recompilation on every call) ---
_RE_NUMBER = re.compile(r'(\d[\d,]*)')
_RE_PROPERTY_PATH = re.compile(r'/Listing/\d|/listing/\d|/property/', re.IGNORECASE)
_RE_VIEW_SPLIT = re.compile(r'(?:View\s+(?:Details|Property|Listing|Photos|OM))\s*·?\s*', re.IGNORECASE)
_RE_BED = re.compile(r'(\d+)[\s-]*(?:bed|licensed\s*bed)', re.IGNORECASE)
_RE_UNIT = re.compile(r'(\d+)[\s-]*unit', re.IGNORECASE)
_RE_CAP = re.compile(r'([\d.]+)\s*%?\s*cap', re.IGNORECASE)
_RE_PRICE = re.compile(r'\$[\d,]+')
_RE_ADDR_ZIP = re.compile(r',\s*[A-Z]{2}\s+\d{5}')
_RE_ADDR_STATE = re.compile(r',\s*[A-Z]{2}\s*$')
_RE_SQFT_CHECK = re.compile(r'[\d,]+\s*(?:sqft|sq\s*ft|sf)\b', re.IGNORECASE)
_RE_SQFT_EXTRACT = re.compile(r'([\d,]+)\s*(?:sqft|sq\s*ft|sf)', re.IGNORECASE)
_RE_PHONE = re.compile(r'\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}')
_RE_EMAIL = re.compile(r'[\w.+-]+@[\w-]+\.[\w.]+')
_RE_STARTS_WITH_DIGIT = re.compile(r'^[\d(+]')
_RE_FULL_ADDR = re.compile(r'(\d+[^,\n]+,\s*[A-Za-z\s]+,\s*[A-Z]{2}\s+\d{5})')
_RE_BED_DETAIL = re.compile(r'(\d+)[\s-]*(?:bed|licensed)', re.IGNORECASE)
_RE_ROOM = re.compile(r'(\d+)[\s-]*room', re.IGNORECASE)


# --- Configuration ---
MIN_BEDS = 50
MIN_UNITS = 40
MAX_PAGES = 12

# States to scrape.
# You can provide either 2-letter postal abbreviations (e.g., "PA") or full state names
# (e.g., "Pennsylvania"). Common misspellings are handled for convenience.
STATES_INPUT = [
    "Alabama",
    "Arkansas",
    "Arizona",
    "California",
    "COlorado",
    "Idaho",
    "Illinois",
    "Maryland",
    "Massachusets",
    "Michigan",
    "Minesota",
    "Montana",
    "Nebraska",
    "Nevada",
    "New Hampshire",
    "NOrth Carolina",
    "Ohio",
    "Okakhoma",
    "Oregon",
    "PA",
    "RI",
    "SC",
    "TN",
    "TX",
    "UT",
    "VT",
    "Wisconsin",
    "WY",
]

def _normalize_state_token(s: str) -> str:
    return re.sub(r"[^a-z]", "", (s or "").lower())

_US_STATE_CODES = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas", "CA": "California",
    "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware", "FL": "Florida", "GA": "Georgia",
    "HI": "Hawaii", "ID": "Idaho", "IL": "Illinois", "IN": "Indiana", "IA": "Iowa",
    "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi", "MO": "Missouri",
    "MT": "Montana", "NE": "Nebraska", "NV": "Nevada", "NH": "New Hampshire", "NJ": "New Jersey",
    "NM": "New Mexico", "NY": "New York", "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio",
    "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah", "VT": "Vermont",
    "VA": "Virginia", "WA": "Washington", "WV": "West Virginia", "WI": "Wisconsin", "WY": "Wyoming",
    "DC": "District of Columbia",
}

_STATE_ALIASES = {}
for _code, _name in _US_STATE_CODES.items():
    _STATE_ALIASES[_normalize_state_token(_code)] = _code
    _STATE_ALIASES[_normalize_state_token(_name)] = _code

# Common misspellings / variants:
_STATE_ALIASES[_normalize_state_token("Massachusets")] = "MA"
_STATE_ALIASES[_normalize_state_token("Minesota")] = "MN"
_STATE_ALIASES[_normalize_state_token("Okakhoma")] = "OK"

def _expand_states(states_input):
    codes = []
    unknown = []
    for raw in states_input:
        key = _normalize_state_token(raw)
        code = _STATE_ALIASES.get(key)
        if not code:
            unknown.append(raw)
            continue
        if code not in codes:
            codes.append(code)
    if unknown:
        raise ValueError(
            "Unknown state(s): "
            + ", ".join(unknown)
            + ". Use a 2-letter code (e.g., 'PA') or full state name (e.g., 'Pennsylvania')."
        )
    return codes

# LoopNet uses lowercase 2-letter state codes in search URLs
STATES = _expand_states(STATES_INPUT)
SEARCH_URLS = [
    f"https://www.loopnet.com/search/assisted-living-facilities/{st.lower()}/for-sale/"
    for st in STATES
]


def get_output_filename():
    """Generate timestamped filename so it never conflicts with an open file."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"loopnet_assisted_living_{ts}.xlsx"


def wait_for_captcha(page, timeout=180):
    warned = False
    start = time.time()
    while time.time() - start < timeout:
        try:
            content = page.content().lower()
        except Exception:
            time.sleep(2)
            continue
        captcha_signs = [
            "verify you are human", "checking your browser", "just a moment",
            "challenge-platform", "turnstile", "hcaptcha", "recaptcha",
            "cf-challenge", "cf-turnstile", "are you a robot",
            "distil_r_captcha", "perimeterx", "px-captcha",
        ]
        if not any(s in content for s in captcha_signs):
            if warned:
                print("  ✅ CAPTCHA cleared!")
            return True
        if not warned:
            print()
            print("  " + "=" * 48)
            print("  ⚠️  CAPTCHA detected!")
            print("  ➡️  Solve it in the browser window.")
            print("  ➡️  Script continues automatically.")
            print(f"  ⏱️  Waiting up to {timeout}s...")
            print("  " + "=" * 48)
            print()
            warned = True
        time.sleep(2)
    print("  ⏰ CAPTCHA timeout.")
    return False


def navigate(page, url):
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
    except PWTimeout:
        print(f"  Page load timeout, continuing...")
    try:
        page.wait_for_load_state("networkidle", timeout=8000)
    except PWTimeout:
        pass
    wait_for_captcha(page)
    time.sleep(1)


def scroll_page(page, times=5, delay=1):
    for _ in range(times):
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(delay)


def extract_number(text):
    if not text:
        return 0
    m = _RE_NUMBER.search(str(text))
    if m:
        try:
            return int(m.group(1).replace(",", ""))
        except ValueError:
            pass
    return 0


def scrape_search_page(page, state=""):
    listings = []
    try:
        page.wait_for_load_state("networkidle", timeout=6000)
    except PWTimeout:
        pass
    scroll_page(page, times=5, delay=1)

    try:
        body_text = page.inner_text("body")
    except Exception:
        body_text = ""

    # --- Find card elements (LoopNet uses various card/placard patterns) ---
    card_selectors = [
        "[class*='placard']", "[class*='Placard']",
        "article.placard", "article[class*='placard']",
        "[data-testid='property-card']", "[class*='PropertyCard']",
        "[class*='property-card']", "[class*='listing-card']",
        "[class*='ListingCard']", "[class*='search-card']",
        "[class*='result-card']", "[class*='search-result']",
        "article[class*='card']", ".listing-card",
        "[class*='profileCard']", "[class*='profile-card']",
    ]
    cards = []
    used_selector = ""
    for sel in card_selectors:
        try:
            found = page.query_selector_all(sel)
            if len(found) > 1:
                cards = found
                used_selector = sel
                break
        except Exception:
            continue

    # --- Fallback: property links ---
    property_links = []
    if not cards:
        try:
            all_links = page.query_selector_all("a")
            seen = set()
            for a in all_links:
                href = a.get_attribute("href") or ""
                if _RE_PROPERTY_PATH.search(href) and href not in seen:
                    if any(skip in href.lower() for skip in [
                        "senior-housing", "senior-living", "assisted-living",
                        "?types=", "?propertyTypes=", "/search/",
                    ]):
                        continue
                    full_url = href if href.startswith("http") else "https://www.loopnet.com" + href
                    seen.add(href)
                    property_links.append({"element": a, "url": full_url})
            if property_links:
                print(f"  Found {len(property_links)} property links")
        except Exception as e:
            print(f"  Link extraction error: {e}")

    # --- Extract data from cards or fallback ---
    if cards:
        print(f"  Found {len(cards)} cards via: {used_selector}")
        for card in cards:
            try:
                text = card.inner_text() or ""
                if len(text.strip()) < 10:
                    continue
                link = ""
                try:
                    a = card.query_selector("a[href*='/Listing/'], a[href*='/listing/'], a[href*='/property/']")
                    if a:
                        h = a.get_attribute("href") or ""
                        if not any(skip in h.lower() for skip in ["search", "?types="]):
                            link = h if h.startswith("http") else "https://www.loopnet.com" + h
                except Exception:
                    pass
                listing = parse_listing_text(text, link, state)
                if listing:
                    listings.append(listing)
            except Exception:
                continue

    elif property_links:
        for pl in property_links:
            try:
                el = pl["element"]
                text = ""
                try:
                    parent = el.evaluate_handle("el => el.closest('div[class]') || el.parentElement || el")
                    text = parent.as_element().inner_text() if parent.as_element() else el.inner_text()
                except Exception:
                    text = el.inner_text() or ""
                if len(text.strip()) < 5:
                    continue
                listing = parse_listing_text(text, pl["url"], state)
                if listing:
                    listings.append(listing)
            except Exception:
                continue
    else:
        print("  No card elements found. Parsing full page text...")
        save_debug(page, body_text)
        blocks = _RE_VIEW_SPLIT.split(body_text)
        for block in blocks:
            block = block.strip()
            if len(block) > 20 and ("$" in block or "bed" in block.lower() or "unit" in block.lower()):
                listing = parse_listing_text(block, "", state)
                if listing:
                    listings.append(listing)

    seen_urls = set()
    unique = []
    for l in listings:
        key = l.get("listing_url") or l.get("address", "")
        if key and key not in seen_urls:
            seen_urls.add(key)
            unique.append(l)

    print(f"  Parsed {len(unique)} listings from this page")
    return unique


def parse_listing_text(text, link="", state=""):
    lines = [l.strip() for l in text.replace("·", "\n").split("\n") if l.strip()]
    if not lines:
        return None

    address = ""
    price = ""
    beds = 0
    units = 0
    sqft = ""
    prop_type = ""
    broker = ""
    cap_rate = ""
    full_text = text.lower()

    bed_match = _RE_BED.search(full_text)
    if bed_match:
        beds = int(bed_match.group(1))

    unit_match = _RE_UNIT.search(full_text)
    if unit_match:
        units = int(unit_match.group(1))

    room_match = _RE_ROOM.search(full_text)
    if room_match and units == 0:
        units = int(room_match.group(1))

    cap_match = _RE_CAP.search(full_text)
    if cap_match:
        cap_rate = cap_match.group(1) + "% CAP"

    _SKIP_PHRASES = {"view details", "view property", "view listing", "view photos",
                     "save search", "sign up", "log in", "show map", "clear filters",
                     "results per page", "save my search", "see new listings",
                     "create alert", "get alerts", "view om", "view flyer"}
    for line in lines:
        ll = line.lower().strip()
        if any(skip in ll for skip in _SKIP_PHRASES):
            continue
        if _RE_PRICE.search(line) and not price:
            price = line.strip()
        elif "price not disclosed" in ll and not price:
            price = "Price Not Disclosed"
        elif "call for" in ll and "price" in ll and not price:
            price = "Call for Price"
        elif "negotiable" in ll and not price:
            price = "Negotiable"
        elif _RE_ADDR_ZIP.search(line) and not address:
            address = line.strip()
        elif _RE_ADDR_STATE.search(line) and not address:
            address = line.strip()
        elif _RE_SQFT_CHECK.search(ll) and not sqft:
            sf_match = _RE_SQFT_EXTRACT.search(ll)
            if sf_match:
                sqft = sf_match.group(1) + " SF"
        elif any(k in ll for k in ["assisted living", "senior living", "senior housing",
                                     "nursing home", "memory care", "skilled nursing",
                                     "continuing care", "medical care", "health care",
                                     "residential care", "adult care", "group home"]) and not prop_type:
            prop_type = line.strip()
        elif any(k in ll for k in ["marcus", "millichap", "cbre", "cushman", "jll",
                                     "colliers", "newmark", "berkadia", "ad advisors",
                                     "fish commercial", "realty", "advisors group",
                                     "capital", "brokerage", "keller williams",
                                     "coldwell banker", "century 21", "nai", "svn"]) and not broker:
            broker = line.strip()

    if not address and not prop_type:
        for line in lines:
            ll = line.lower()
            if any(skip in ll for skip in ["view", "save", "sign up",
                                             "show map", "clear", "results per"]):
                continue
            if len(line) > 15 and "$" not in line:
                prop_type = line.strip()
                break

    if not price and beds == 0 and units == 0:
        return None

    return {
        "address": address, "price": price, "beds": beds, "units": units,
        "sqft": sqft, "property_type": prop_type, "cap_rate": cap_rate,
        "broker_name": broker, "broker_phone": "", "broker_email": "",
        "broker_company": "", "listing_url": link, "state": state,
    }


def scrape_detail_page(page, url):
    info = {"broker_name": "", "broker_phone": "", "broker_email": "",
            "broker_company": "", "beds": 0, "units": 0, "sqft": "", "address": "",
            "cap_rate": "", "property_type": ""}
    try:
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=20000)
        except PWTimeout:
            pass
        try:
            page.wait_for_load_state("networkidle", timeout=5000)
        except PWTimeout:
            pass
        wait_for_captcha(page, timeout=60)
        scroll_page(page, times=2, delay=0.8)

        body_text = page.inner_text("body") or ""

        # Try to find broker/contact section
        contact_text = ""
        for sel in ["[class*='roker']", "[class*='ontact']", "[class*='Agent']",
                     "[class*='Team']", "[class*='Advisor']", "[class*='advisor']",
                     "[class*='listing-broker']", "[class*='ListingBroker']",
                     "[class*='broker-card']", "[class*='BrokerCard']",
                     "[class*='contact-card']", "[class*='ContactCard']"]:
            try:
                for el in page.query_selector_all(sel):
                    t = el.inner_text() or ""
                    if t.strip():
                        contact_text += t + "\n"
            except Exception:
                continue

        search_text = contact_text if contact_text else body_text

        # Phone from tel: links
        try:
            for el in page.query_selector_all("a[href^='tel:']"):
                href = el.get_attribute("href") or ""
                phone = href.replace("tel:", "").strip()
                if phone and len(phone) >= 10:
                    info["broker_phone"] = phone
                    break
        except Exception:
            pass
        if not info["broker_phone"]:
            phones = _RE_PHONE.findall(search_text)
            if phones:
                info["broker_phone"] = phones[0]

        # Email from mailto: links
        try:
            for el in page.query_selector_all("a[href^='mailto:']"):
                href = el.get_attribute("href") or ""
                email = href.replace("mailto:", "").split("?")[0].strip()
                if email and "@" in email and "loopnet" not in email.lower():
                    info["broker_email"] = email
                    break
        except Exception:
            pass
        if not info["broker_email"]:
            for e in _RE_EMAIL.findall(search_text):
                if "loopnet.com" not in e.lower() and "costar.com" not in e.lower():
                    info["broker_email"] = e
                    break

        # Broker name
        if contact_text:
            for line in contact_text.split("\n"):
                line = line.strip()
                if not line or "@" in line or _RE_STARTS_WITH_DIGIT.match(line) or len(line) > 50 or len(line) < 4:
                    continue
                if any(k in line.lower() for k in ["contact", "listed", "team", "view",
                                                     "request", "schedule", "call", "share",
                                                     "save", "print", "report", "broker",
                                                     "question", "interest", "tour",
                                                     "message", "inquiry"]):
                    continue
                words = line.split()
                if 2 <= len(words) <= 4 and all(w[0].isupper() for w in words if len(w) > 1 and w[0].isalpha()):
                    info["broker_name"] = line
                    break

        # Broker company
        for line in search_text.split("\n"):
            ll = line.lower().strip()
            if any(k in ll for k in ["marcus", "millichap", "cbre", "cushman", "jll",
                                       "colliers", "newmark", "berkadia", "ad advisors",
                                       "fish commercial", "keller williams",
                                       "coldwell banker", "century 21", "nai", "svn"]):
                info["broker_company"] = line.strip()
                break

        # Beds / Units / SF (cache lowered text to avoid repeated .lower())
        body_lower = body_text.lower()
        bed_match = _RE_BED_DETAIL.search(body_lower)
        if bed_match:
            info["beds"] = int(bed_match.group(1))
        unit_match = _RE_UNIT.search(body_lower)
        if unit_match:
            info["units"] = int(unit_match.group(1))
        if not info["units"]:
            room_match = _RE_ROOM.search(body_lower)
            if room_match:
                info["units"] = int(room_match.group(1))
        sf_match = _RE_SQFT_EXTRACT.search(body_lower)
        if sf_match:
            info["sqft"] = sf_match.group(1) + " SF"

        # Address
        addr_match = _RE_FULL_ADDR.search(body_text)
        if addr_match:
            info["address"] = addr_match.group(1).strip()

        cap_match = _RE_CAP.search(body_lower)
        if cap_match:
            info["cap_rate"] = cap_match.group(1) + "% CAP"

    except Exception as e:
        print(f"    Detail error: {e}")
    return info


def save_debug(page, body_text=""):
    try:
        with open("debug_loopnet_page.html", "w", encoding="utf-8") as f:
            f.write(page.content())
        if not body_text:
            body_text = page.inner_text("body")
        with open("debug_loopnet_text.txt", "w", encoding="utf-8") as f:
            f.write(body_text)
        print("  📄 Saved debug_loopnet_page.html and debug_loopnet_text.txt")
    except Exception:
        pass


def click_next(page):
    # Try targeted CSS selectors first (much faster than iterating all elements)
    for sel in ["[aria-label='Next']", "[aria-label='next']", "[aria-label='Next Page']",
                "[class*='next' i]", "[class*='Next']",
                "a[class*='paging-next']", "a[class*='pagingNext']",
                "[class*='pagination'] a:last-child"]:
        try:
            btn = page.query_selector(sel)
            if btn and btn.is_visible():
                btn.click()
                try:
                    page.wait_for_load_state("networkidle", timeout=5000)
                except PWTimeout:
                    time.sleep(2.5)
                return True
        except Exception:
            continue
    # Fallback: iterate all links/buttons looking for "Next" text
    try:
        for el in page.query_selector_all("a, button"):
            txt = (el.inner_text() or "").strip()
            if txt in ["Next", "›", "»", ">", "Next Page"]:
                if el.is_visible() and el.is_enabled():
                    el.click()
                    try:
                        page.wait_for_load_state("networkidle", timeout=5000)
                    except PWTimeout:
                        time.sleep(2.5)
                    return True
    except Exception:
        pass
    return False


def meets_criteria(lst):
    beds = lst.get("beds", 0) or 0
    units = lst.get("units", 0) or 0
    if beds == 0 and units == 0:
        return True
    return beds >= MIN_BEDS or units >= MIN_UNITS


def dedupe(listings):
    seen = set()
    out = []
    for l in listings:
        key = l.get("listing_url", "") or l.get("address", "").strip().lower() or l.get("price", "")
        if key and key not in seen:
            seen.add(key)
            out.append(l)
    return out


def write_excel(listings, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Assisted Living - LoopNet"

    hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hfl = PatternFill("solid", fgColor="2F5496")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    df = Font(name="Arial", size=10)
    da = Alignment(vertical="top", wrap_text=True)
    lf = Font(name="Arial", size=10, color="0563C1", underline="single")
    bd = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
    af = PatternFill("solid", fgColor="F2F7FB")

    ws.merge_cells("A1:M1")
    ws["A1"] = f"LoopNet — Healthcare / Assisted Living For Sale ({', '.join(STATES)}) — 50+ Beds / 40+ Units"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:M2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}  |  Source: loopnet.com"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")

    hdrs = ["#", "State", "Property Address", "Asking Price", "CAP Rate", "Beds", "Units",
            "Sq Ft", "Property Type", "Broker / Company", "Broker Phone", "Broker Email", "Listing URL"]
    wds = [5, 7, 40, 16, 10, 8, 8, 14, 24, 26, 18, 28, 50]
    hr = 4
    for ci, (h, w) in enumerate(zip(hdrs, wds), 1):
        c = ws.cell(row=hr, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfl, ha, bd
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[hr].height = 28

    for i, lst in enumerate(listings, 1):
        row = hr + i
        fill = af if i % 2 == 0 else PatternFill()
        broker_co = lst.get("broker_name", "")
        if lst.get("broker_company"):
            broker_co = lst["broker_company"]
            if lst.get("broker_name"):
                broker_co = f"{lst['broker_name']} — {lst['broker_company']}"

        vals = [i, lst.get("state", ""), lst.get("address", ""), lst.get("price", ""),
                lst.get("cap_rate", ""),
                lst.get("beds") or "", lst.get("units") or "", lst.get("sqft", ""),
                lst.get("property_type", ""), broker_co,
                lst.get("broker_phone", ""), lst.get("broker_email", ""),
                lst.get("listing_url", "")]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = lf if ci == 13 and v else df
            c.alignment = da
            c.border = bd
            if fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
                c.fill = fill
        if lst.get("listing_url"):
            ws.cell(row=row, column=13).hyperlink = lst["listing_url"]

    sr = hr + len(listings) + 2
    ws.merge_cells(f"A{sr}:M{sr}")
    ws.cell(row=sr, column=1, value=f"Total: {len(listings)} listings").font = Font(
        name="Arial", bold=True, size=10, color="2F5496")
    ws.freeze_panes = f"A{hr + 1}"
    ws.auto_filter.ref = f"A{hr}:M{hr + max(len(listings), 1)}"

    try:
        wb.save(filename)
        print(f"\n✅ Saved: {filename} ({len(listings)} listings)")
    except PermissionError:
        alt = filename.replace(".xlsx", "_v2.xlsx")
        wb.save(alt)
        print(f"\n✅ Original file was locked. Saved as: {alt} ({len(listings)} listings)")


def main():
    print("=" * 60)
    print("  LoopNet Assisted Living Scraper v2 (Playwright)")
    print(f"  Category: Healthcare > Assisted Living")
    print(f"  States: {', '.join(STATES)}")
    print(f"  Filter: {MIN_BEDS}+ beds OR {MIN_UNITS}+ units")
    print("=" * 60)
    print()
    print("  A browser will open. If CAPTCHA appears,")
    print("  solve it manually — script auto-continues.")
    print()

    output_file = get_output_filename()
    print(f"  Output file: {output_file}")
    print()

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        )
        context.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        pg = context.new_page()

        all_listings = []

        try:
            print("Opening LoopNet...")
            navigate(pg, "https://www.loopnet.com")

            for i, url in enumerate(SEARCH_URLS):
                state = STATES[i]
                print(f"\n🔍 [{state}] {url}")
                navigate(pg, url)

                if i == 0:
                    save_debug(pg)

                page_num = 1
                while page_num <= MAX_PAGES:
                    print(f"  📄 Page {page_num}...")
                    cards = scrape_search_page(pg, state)
                    if not cards:
                        print("  No more listings.")
                        break
                    all_listings.extend(cards)
                    if not click_next(pg):
                        print("  No next page.")
                        break
                    page_num += 1
                    wait_for_captcha(pg, timeout=30)

            all_listings = dedupe(all_listings)
            print(f"\n📋 Total unique listings: {len(all_listings)}")

            if all_listings:
                print("\n📞 Getting broker details...")
                for i, lst in enumerate(all_listings):
                    url = lst.get("listing_url", "")
                    if not url:
                        continue
                    # Skip detail scrape if we already have broker contact info
                    if lst.get("broker_phone") and lst.get("broker_email") and lst.get("broker_name"):
                        print(f"  [{i+1}/{len(all_listings)}] Skipping (already have broker info)")
                        continue
                    print(f"  [{i+1}/{len(all_listings)}] {url[:70]}...")
                    detail = scrape_detail_page(pg, url)
                    for k in detail:
                        if detail[k] and not lst.get(k):
                            lst[k] = detail[k]
                    time.sleep(0.75)

            filtered = [l for l in all_listings if meets_criteria(l)]
            final = []
            for l in filtered:
                beds = l.get("beds", 0) or 0
                units = l.get("units", 0) or 0
                if beds > 0 and beds < MIN_BEDS and units > 0 and units < MIN_UNITS:
                    continue
                final.append(l)

            print(f"\n📊 After size filter: {len(final)} listings")
            write_excel(final, output_file)

            if not final:
                print("\n⚠️  No listings found. Check debug files.")

        finally:
            input("\nPress Enter to close browser...")
            browser.close()
            print("Done!")


if __name__ == "__main__":
    main()
